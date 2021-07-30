import openpyxl

path = 'D:\\Обработка файлов таможни\\Пример.xlsx' # начинаем сортировку по производителю
wb = openpyxl.load_workbook(path)
sheet = wb['2019']
quantity_row = sheet.max_row
quantity_column = sheet.max_column
data = []
row_i = []
for i in range(2, quantity_row + 1):
    for j in range(1, quantity_column + 1):
        a = sheet.cell(row=i, column=j)
        b = str(a.value)
        row_i.append(b)
    row_i = tuple(row_i)
    data.append(row_i)
    row_i = []
data.sort(key=lambda product: product[1])

for i in range(0, len(data)):
    for j in range(0, len(data[0])):
        val = data[i][j]
        sheet.cell(row=i + 2, column=j + 1).value = val

wb.save('D:\\Обработка файлов таможни\\Пример_1.xlsx')
