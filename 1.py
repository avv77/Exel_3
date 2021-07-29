import openpyxl

path = 'D:\\Работа\\Химия\\Хлорид железа\\Импорт_хлорид железа_1.xlsx'
wb = openpyxl.load_workbook(path)
sheet = wb['2019']
quantity_row = sheet.max_row
quantity_column = sheet.max_column
for i in range(1, quantity_row + 1):
    for j in range(1, quantity_column + 1):
        a = sheet.cell(row=i, column=j)
        b = str(a.value)
        for k in b:
            if k == '`':
                c = b.replace('`', '')
                d = sheet.cell(row=i, column=j, value=c)

wb.create_sheet('Prob')

path_1 = 'D:\\Работа\\1_1.xlsx'
wb_1 = openpyxl.load_workbook(path_1)
sheet_1 = wb_1.active
quantity_row_1 = sheet_1.max_row
quantity_column_1 = sheet_1.max_column
for i in range(1, quantity_row_1 + 1):
    for j in range(1, quantity_column_1 + 1):
        a = sheet_1.cell(row=i, column=j)
        b = a.value
        sheet2 = wb['Prob']
        c = sheet2.cell(row=i, column=j, value=b)


sheet = wb['2019']
sheet.insert_cols(idx=76)  # добавляем столбец "Страны"
sheet['BX1'] = 'Страны'
quantity_row = sheet.max_row  # считаем количество строк
for i in range(2, quantity_row + 1):
    formula = '=VLOOKUP(BW2, Prob!A:B, 2, 0)'
    d = sheet.cell(row=i, column=76, value=formula)

wb.save('D:\\Работа\\Химия\\Хлорид железа\\Импорт_хлорид железа1_1.xlsx')
